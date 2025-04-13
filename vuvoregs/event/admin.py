"""Admin configuration for the event management application.

This module defines custom admin interfaces, forms, and actions for managing
events, athletes, registrations, payments, and related entities in the Django
admin panel.
"""

# 🔧 Core Python imports
import csv
import io
import json

from django import forms
from django.conf import settings

# 🧩 Django core admin functionality
from django.contrib import admin, messages
from django.db import transaction

# 🌐 HTTP, routing, and utilities
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.test import RequestFactory
from django.urls import path, reverse
from django.utils.html import format_html

# 🧱 Third-party widgets
from django_json_widget.widgets import JSONEditorWidget
from openpyxl import load_workbook

from .excel import generate_event_template
from .forms import BibNumberImportForm, ExportEventAthletesForm, TeamExcelUploadForm

# 📦 Local app models and forms
from .models.athlete import Athlete
from .models.event import Event, PickUpPoint
from .models.package import RacePackage, RaceSpecialPrice, PackageOption
from .models.payment import Payment
from .models.race import Race, RaceType, TimeBasedPrice
from .models.registration import Registration
from .models.terms import TermsAndConditions


# 👤 Custom form with JSON widget
class AthleteAdminForm(forms.ModelForm):
    class Meta:
        model = Athlete
        fields = [
            "first_name",
            "last_name",
            "email",
            "phone",
            "package",
            "race",
            "pickup_point",
            "selected_options",
        ]
        widgets = {
            "selected_options": JSONEditorWidget(
                options={"mode": "form", "mainMenuBar": True}
            ),
        }


# 👥 Inline for displaying athletes under registrations
class AthleteInline(admin.TabularInline):
    model = Athlete
    extra = 0
    readonly_fields = ["formatted_selected_options"]
    fields = [
        "first_name",
        "last_name",
        "email",
        "package",
        "race",
        "formatted_selected_options",
    ]

    def formatted_selected_options(self, obj):
        if not obj.selected_options:
            return "-"
        try:
            return "\n".join(
                f"{k}: {', '.join(v)}" for k, v in obj.selected_options.items()
            )
        except Exception:
            return "⚠️ Invalid JSON"

    formatted_selected_options.short_description = "Selected Options"


@admin.register(TermsAndConditions)
class TermsAndConditionsAdmin(admin.ModelAdmin):
    list_display = ("event", "version", "created_at")
    search_fields = ("event__name", "version", "title")
    ordering = ("-created_at",)


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ("name", "date", "location", "is_available")
    list_filter = ("is_available",)
    search_fields = ("name", "location")
    ordering = ("-date",)
    change_form_template = "admin/event/change_form_with_download.html"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "<int:event_id>/download-template/",
                self.admin_site.admin_view(self.download_template),
                name="event_download_template",
            ),
            path(
                "<int:event_id>/upload-teams/",
                self.admin_site.admin_view(self.upload_team_excel),
                name="event_upload_teams",
            ),
        ]
        return custom + urls

    def download_template(self, request, event_id):
        event = get_object_or_404(Event, pk=event_id)
        return generate_event_template(event)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        extra_context = extra_context or {}
        extra_context["download_template_url"] = reverse(
            "admin:event_download_template", args=[object_id]
        )
        extra_context["upload_excel_url"] = reverse(
            "admin:event_upload_teams", args=[object_id]
        )
        return super().change_view(
            request, object_id, form_url, extra_context=extra_context
        )

    def upload_team_excel(self, request, *args, **kwargs):
        event_id = kwargs.get("event_id")
        event = get_object_or_404(Event, pk=event_id)

        context = {
            "form": TeamExcelUploadForm(),
            "event": event,
            "title": f"Team Upload for {event.name}",
            **self.admin_site.each_context(request),
        }

        if request.method == "POST":
            form = TeamExcelUploadForm(request.POST, request.FILES)
            context["form"] = form
            error_log = []

            if form.is_valid():
                wb = load_workbook(form.cleaned_data["excel_file"])
                ws = wb.active
                headers = [cell.value for cell in ws[1]]

                validated_rows = []
                row_index = 2

                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row, strict=False))
                    row_has_error = False

                    if not data.get("first_name") or not data.get("last_name"):
                        error_log.append(
                            f"Row {row_index}: Missing first_name or last_name."
                        )
                        row_index += 1
                        continue

                    try:
                        race = Race.objects.get(name=data["race_name"], event=event)
                    except Race.DoesNotExist:
                        error_log.append(
                            f"Row {row_index}: Race '{data.get('race_name')}' not found."
                        )
                        row_index += 1
                        continue

                    try:
                        package = RacePackage.objects.get(
                            name=data["package_name"], event=event, races=race
                        )
                    except RacePackage.DoesNotExist:
                        error_log.append(
                            f"Row {row_index}: Package '{data.get('package_name')}' not found for race '{race.name}'."
                        )
                        row_index += 1
                        continue

                    selected_options = {}
                    package_options = {
                        opt.name: opt for opt in package.packageoption_set.all()
                    }

                    for k, v in data.items():
                        if k and str(k).startswith("option_") and v:
                            opt_name = k.replace("option_", "").strip()
                            value = str(v).strip()

                            option = package_options.get(opt_name)
                            if not option:
                                error_log.append(
                                    f"Row {row_index}: Unknown option '{opt_name}' for package '{package.name}'."
                                )
                                row_has_error = True
                                break

                            valid_choices = option.options_json
                            match = next(
                                (
                                    c
                                    for c in valid_choices
                                    if c.lower() == value.lower()
                                ),
                                None,
                            )

                            if not match:
                                error_log.append(
                                    f"Row {row_index}: Invalid value '{value}' for option '{opt_name}'. "
                                    f"Valid options: {', '.join(valid_choices)}"
                                )
                                row_has_error = True
                                break

                            selected_options[opt_name] = match

                    for opt_name, opt in package_options.items():
                        if opt.options_json and opt_name not in selected_options:
                            # Only flag as missing if it wasn’t already attempted (no invalid attempt earlier)
                            provided_keys = [
                                k.replace("option_", "").strip()
                                for k in data.keys()
                                if k and str(k).startswith("option_")
                            ]
                            if opt_name not in provided_keys:
                                error_log.append(
                                    f"Row {row_index}: Missing required option '{opt_name}' for package '{package.name}'."
                                )
                                row_has_error = True
                                break

                    if not row_has_error:
                        validated_rows.append((data, race, package, selected_options))

                    row_index += 1

                if error_log:
                    context["error_log"] = error_log
                    return render(request, "admin/team_excel_upload.html", context)

                with transaction.atomic():
                    reg = Registration.objects.create(event=event)
                    for data, race, package, selected_options in validated_rows:
                        Athlete.objects.create(
                            registration=reg,
                            race=race,
                            package=package,
                            first_name=data.get("first_name", "").strip(),
                            last_name=data.get("last_name", "").strip(),
                            email=data.get("email", ""),
                            phone=data.get("phone", ""),
                            dob=data.get("dob"),
                            sex=str(data.get("sex", "")).upper(),
                            hometown=data.get("hometown", ""),
                            team=data.get("team", ""),
                            selected_options=selected_options,
                        )
                    reg.total_amount = sum(
                        a.package.get_current_final_price() for a in reg.athletes.all()
                    )
                    reg.save()

                messages.success(
                    request, f"✅ Successfully uploaded {len(validated_rows)} athletes."
                )
                return redirect("admin:event_event_changelist")

        return render(request, "admin/team_excel_upload.html", context)

    @admin.register(RaceType)
    class RaceTypeAdmin(admin.ModelAdmin):
        list_display = ("name",)
        search_fields = ("name",)

    @admin.register(Race)
    class RaceAdmin(admin.ModelAdmin):
        list_display = (
            "name",
            "race_type",
            "race_km",
            "event",
            "base_price_individual",
            "base_price_team",
        )
        list_filter = ("event", "race_type")
        search_fields = ("race_type__name", "event__name")
        ordering = ("event", "race_type")
        fields = (
            "event",
            "name",
            "race_type",
            "race_km",
            "max_participants",
            "min_participants",
            "team_discount_threshold",
            "base_price_individual",
            "base_price_team",
            "image",
        )


class RacePackageOptionInline(admin.TabularInline):
    model = PackageOption
    extra = 1


@admin.register(RacePackage)
class RacePackageAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "event",
        "price_adjustment",
        "visible_until",
    )
    list_filter = ("event",)
    search_fields = ("name", "event__name")
    ordering = ("event", "name")
    inlines = [RacePackageOptionInline]


@admin.register(Registration)
class RegistrationAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "event",
        "created_at",
        "status",
        "payment_status",
        "total_amount",
        "num_athletes",
        "get_terms_version",
        "agrees_to_terms",
        "payment_link",
    )
    list_filter = ("event", "status", "payment_status", "agrees_to_terms")
    search_fields = ("id", "event__name")
    ordering = ("-created_at",)
    readonly_fields = (
        "created_at",
        "updated_at",
        "total_amount",
        "get_terms_version",
        "num_athletes",
        "payment_link",
    )
    inlines = [AthleteInline]

    @admin.display(description="T&Cs Version")
    def get_terms_version(self, obj):
        return obj.agreed_to_terms.version if obj.agreed_to_terms else "—"

    @admin.display(description="Athletes")
    def num_athletes(self, obj):
        return obj.athletes.count()

    @admin.display(description="Payment")
    def payment_link(self, obj):
        if obj.payment:
            url = reverse("admin:event_payment_change", args=[obj.payment.id])
            return format_html(
                '<a href="{}">#{} – {}</a>', url, obj.payment.id, obj.payment.status
            )
        return "—"


@admin.register(PickUpPoint)
class PickUpPointAdmin(admin.ModelAdmin):
    list_display = ("name", "event", "address", "working_hours")
    list_filter = ("event",)
    search_fields = ("name", "address", "event__name")


@admin.action(description="📤 Export selected athletes to CSV")
def export_athletes_to_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=athletes_export.csv"
    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "id",
        "first_name",
        "last_name",
        "email",
        "phone",
        "package",
        "race",
        "pickup_point",  # ✅ Added safely
        "bib_number",
    ])
    for athlete in queryset:
        writer.writerow([
            athlete.id,
            athlete.first_name,
            athlete.last_name,
            athlete.email,
            athlete.phone,
            athlete.package.name if athlete.package else "",
            athlete.race.name if athlete.race else "",
            athlete.pickup_point.name if athlete.pickup_point else "",  # ✅ Null-safe
            athlete.bib_number or "",
        ])
    return response


@admin.register(Athlete)
class AthleteAdmin(admin.ModelAdmin):
    form = AthleteAdminForm
    ordering = ["-registration__created_at"]
    list_display = (
        "registration__created_at",
        "first_name",
        "last_name",
        "email",
        "race",
        "package",
        "pickup_point",
        "dob",
        "special_price",
        "formatted_selected_options",
    )
    list_filter = ("race__event", "race", "package", "pickup_point")
    search_fields = ("first_name", "last_name", "race__name", "email")
    readonly_fields = ["formatted_selected_options"]
    actions = [export_athletes_to_csv]
    fields = (
        "first_name",
        "last_name",
        "dob",
        "sex",
        "email",
        "phone",
        "pickup_point",
        "race",
        "package",
        "registration",
        "selected_options",
    )

    def get_status(self, obj):
        return obj.registration.status if obj.registration else "-"

    get_status.short_description = "Registration Status"

    def formatted_selected_options(self, obj):
        if not obj.selected_options:
            return "-"
        try:
            return "\n".join(
                f"{k}: {', '.join(v)}" for k, v in obj.selected_options.items()
            )
        except Exception:
            return "⚠️ Invalid JSON"

    formatted_selected_options.short_description = "Package Options"


# ➕ Bonus: Add views manually (optional, requires URL config)
def import_bibs_view(request):
    form = BibNumberImportForm()
    success, failed = 0, 0
    if request.method == "POST":
        form = BibNumberImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["csv_file"]
            reader = csv.DictReader(
                io.TextIOWrapper(file, encoding="utf-8"), delimiter=";"
            )
            for row in reader:
                athlete_id = row.get("id")
                bib = row.get("bib_number")
                try:
                    athlete = Athlete.objects.get(id=athlete_id)
                    athlete.bib_number = bib
                    athlete.save()
                    success += 1
                except Athlete.DoesNotExist:
                    failed += 1
            messages.success(request, f"{success} bib numbers updated.")
            if failed:
                messages.warning(
                    request, f"{failed} rows failed (no matching Athlete ID)."
                )
            return redirect("admin:index")
    return render(
        request, "admin/import_bibs.html", {"form": form, "title": "Import Bib Numbers"}
    )


def export_athletes_view(request):
    form = ExportEventAthletesForm()
    if request.method == "POST":
        form = ExportEventAthletesForm(request.POST)
        if form.is_valid():
            event = form.cleaned_data["event"]
            athletes = Athlete.objects.filter(race__event=event).select_related(
                "package", "race"
            )
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = (
                f"attachment; filename=athletes_{event.name or event.id}.csv"
            )
            writer = csv.writer(response, delimiter=";")
            writer.writerow([
                "id",
                "first_name",
                "last_name",
                "dob",
                "package",
                "race",
                "pickup_point",
                "bib_number",
            ])
            for athlete in athletes:
                writer.writerow([
                    athlete.id,
                    athlete.first_name,
                    athlete.last_name,
                    athlete.dob,
                    athlete.pickup_point.name if athlete.pickup_point else "",
                    athlete.package.name if athlete.package else "",
                    athlete.race.name if athlete.race else "",
                    athlete.bib_number or "",
                ])
            return response
    return render(
        request,
        "admin/export_athletes.html",
        {"form": form, "title": "Export Event Athletes"},
    )


@admin.action(description="Set payment status to 'confirmed'")
def simulate_success(modeladmin, request, queryset):
    for payment in queryset:
        payment.status = "confirmed"
        payment.save()
        if hasattr(payment, "registration"):
            registration = payment.registration
            registration.status = "completed"
            registration.payment_status = "paid"
            registration.save()


@admin.action(description="Set payment status to 'failed'")
def simulate_failure(modeladmin, request, queryset):
    for payment in queryset:
        payment.status = "rejected"
        payment.save()
        if hasattr(payment, "registration"):
            registration = payment.registration
            registration.status = "failed"
            registration.payment_status = "failed"
            registration.save()


@admin.action(description="🚀 Simulate Webhook for selected payments")
def simulate_webhook(modeladmin, request, queryset):
    from .views import payment_webhook

    # ✅ Only allow if DEBUG=True
    if not settings.DEBUG:
        messages.error(request, "❌ This action is only available in development mode.")
        return

    factory = RequestFactory()

    for payment in queryset:
        data = json.dumps({"payment_id": str(payment.id)})
        simulated_request = factory.post(
            reverse("payment_webhook"), data=data, content_type="application/json"
        )
        response = payment_webhook(simulated_request)
        if isinstance(response, JsonResponse) and response.status_code == 200:
            messages.success(request, f"✅ Webhook simulated for payment #{payment.id}")
        else:
            messages.error(request, f"❌ Webhook failed for payment #{payment.id}")


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    """Admin interface for managing Payment objects.

    This class provides functionality to display, filter, search, and perform
    actions on Payment records in the Django admin interface.
    """

    list_display = (
        "id",
        "variant",
        "status",
        "total",
        "order_code",
        "currency",
        "billing_email",
        "created",
        "modified",
    )
    list_filter = ("variant", "status", "currency")
    search_fields = ("billing_email", "id", "extra_data")
    readonly_fields = ("created", "modified", "captured_amount")
    actions = [simulate_success, simulate_failure, simulate_webhook]


class RaceSpecialPriceInline(admin.TabularInline):
    model = RaceSpecialPrice
    extra = 1


class TimeBasedPriceInline(admin.TabularInline):
    model = TimeBasedPrice
    extra = 1


# iwill refactor later


class RaceAdmin(admin.ModelAdmin):
    inlines = [TimeBasedPriceInline, RaceSpecialPriceInline]


admin.site.unregister(Race)  # If already registered
admin.site.register(Race, RaceAdmin)
