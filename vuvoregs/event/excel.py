from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse


def generate_event_template(event):
    """
    Generate a dynamic Excel template for team registration for the given event,
    using race_name and package_name instead of IDs.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Upload Template"

    # Standard athlete fields
    headers = [
        "race_name",
        "package_name",
        "first_name",
        "last_name",
        "email",
        "phone",
        "dob",
        "sex",
        "hometown",
        "team",
    ]

    # Dynamic option fields
    option_headers = []
    for race in event.races.all().prefetch_related("packages__packageoption_set"):
        for package in race.packages.all():
            for option in package.packageoption_set.all():
                col = f"option_{option.name}"
                if col not in option_headers:
                    option_headers.append(col)

    headers.extend(option_headers)

    # Write headers to row 1
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)

    # Write sample rows (race + package pairs)
    row = 2
    for race in event.races.all():
        for package in race.packages.all():
            ws.cell(row=row, column=1, value=race.name)
            ws.cell(row=row, column=2, value=package.name)
            row += 1

    # Format column widths
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 2, 15)

    # Export to Excel file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{event.name.lower().replace(' ', '_')}_team_upload_template.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
